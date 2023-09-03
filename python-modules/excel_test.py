from openpyxl import load_workbook


def excel_test(path: str) -> int:
    workbook = load_workbook(path)
    return workbook['Sheet1'].max_row
